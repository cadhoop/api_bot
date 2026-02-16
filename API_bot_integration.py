from flask import Flask, request, jsonify
from datetime import datetime
import mysql.connector
from mysql.connector import Error
from typing import Dict, Any, List, Optional
import os
from datetime import date, datetime
from typing import Dict, List, Tuple
import re
import logging
import sys
import time
import json
import copy
from functools import wraps
import socket
import unicodedata
import hashlib
import hmac
import pandas as pd
from decimal import Decimal
from openpyxl.utils import get_column_letter
import math
import argparse
import psutil
import secrets
import threading
import paramiko
from pathlib import Path
from fpdf import FPDF
from PIL import Image
from fpdf.enums import XPos, YPos
import traceback



from API_bot_parameters_integration import DB_CONFIG_ikoula3, DB_CONFIG_ekima, FIELD_MAPPING, TAB_STOPWORDS, TAB_MOTS_SUP100000_SANS_ACCENT, MOIS_ANNEE, TABLE_FAST,TABLE_ALL,TABLE_AFNIC,MIN_FULLTEXT_LENGTH,LIMIT_DISPLAY_INFO,UNITARY_PRICE_LEGAL_INFOS,MAX_DAILY_REQUESTS_NUMBER, FRENCH_ELISIONS,AUTH_WINDOW , MESSAGE_WORD_TOO_COMMON_1, MESSAGE_WORD_TOO_COMMON_2, MEMORY_THRESHOLD_PERCENT, PATH_REMOTE_OVH, LINK_REMOTE_OVH, SFTP_NAME,SFTP_PORT,SFTP_USERNAME,SFTP_PASSWORD,PATH_REMOTE_DATA_FILE,PATH_REMOTE_INVOICE_FILE,PATH_LOCAL_INVOICE_FILE,PATH_LOCAL_DATA_FILE, PATH_LOGO_MARKETHINGS_EKIMIA, PATH_LOGO_MARKETHINGS_IKOULA3


app = Flask(__name__)
app.config['JSON_AS_ASCII'] = False

# Empêche Flask de trier les clés alphabétiquement
app.json.sort_keys = False

# nohup python3 API_bot_integration.py --verbose="no" > app.log 2>&1 &


# 1. Initialiser le parseur
parser = argparse.ArgumentParser(description="Script API Bot avec mode verbose")

# 2. Définir l'argument --verbose
# On utilise default="no" pour que le script fonctionne même sans l'argument
parser.add_argument('--verbose', type=str, default="no", help='Activer le mode verbose (yes/no)')

# 3. Récupérer les arguments
args = parser.parse_args()

VERBOSE = False
# 4. Utilisation dans votre code
if args.verbose == "yes":
    print("--- MODE VERBOSE ACTIVÉ ---")
    print(f"Arguments reçus : {args}")
    VERBOSE = True
    
FPDF_VERSION = 2


hostname = socket.gethostname()
if hostname == "frhb96148ds":
    DB_CONFIG = DB_CONFIG_ikoula3
    PATH_LOGO_MARKETHINGS = PATH_LOGO_MARKETHINGS_IKOULA3

elif hostname == "dhoop-NS5x-NS7xAU":
    DB_CONFIG = DB_CONFIG_ekima
    PATH_LOGO_MARKETHINGS = PATH_LOGO_MARKETHINGS_EKIMIA


logger = logging.getLogger()
logger.setLevel(logging.DEBUG)  # keep DEBUG for file logging

# Remove default handlers added by basicConfig (if any)
if logger.hasHandlers():
    logger.handlers.clear()

# Console handler (WARNING+ only)
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setLevel(logging.WARNING)  # show only WARNING, ERROR, CRITICAL
console_formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')
console_handler.setFormatter(console_formatter)
logger.addHandler(console_handler)

# File handler (DEBUG+ everything)
file_handler = logging.FileHandler("api.log")
file_handler.setLevel(logging.DEBUG)
file_formatter = logging.Formatter('%(asctime)s [%(levelname)s] %(message)s')
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)

# Suppress Flask werkzeug INFO logs in console
werkzeug_logger = logging.getLogger('werkzeug')
werkzeug_logger.setLevel(logging.WARNING)

logger = logging.getLogger(__name__)

# Load API_KEYS from env defense code
try:
    raw = os.getenv("API_KEYS", "{}")

    if VERBOSE:
        print(f"*********************raw:{raw}")
    API_KEYS_V1 = json.loads(raw)
    if VERBOSE:
        print(f"API_KEYS_V1:{API_KEYS_V1}")
    flag_error_key_V1 = False

except json.JSONDecodeError:
    flag_error_key_V1 = True
    raise RuntimeError("API_KEYS_V1 environment variable must be valid JSON")

try:
    raw = os.getenv("API_KEYS_V2", "{}")
    API_KEYS_V2 = json.loads(raw)
    flag_error_key_V2 = False

except json.JSONDecodeError:
    flag_error_key_V2 = True
    raise RuntimeError("API_KEYS_V2 environment variable must be valid JSON")

if (flag_error_key_V2 or flag_error_key_V1):
    print("error lecture des clefs")
    sys.exit();
else:
    print("Chargement OK des clefs d'authentification")

# Defensive fix: unwrap list if accidentally wrapped
if isinstance(API_KEYS_V1, list) and len(API_KEYS_V1) == 1 and isinstance(API_KEYS_V1[0], str):
    try:
        API_KEYS_V1 = json.loads(API_KEYS_V1[0])
    except json.JSONDecodeError:
        raise RuntimeError("API_KEYS_V1 list contains invalid JSON")

if not isinstance(API_KEYS_V1, dict):
    raise RuntimeError("API_KEYS_V1 must be a JSON object (dict)")

# Defensive fix: unwrap list if accidentally wrapped
if isinstance(API_KEYS_V2, list) and len(API_KEYS_V2) == 1 and isinstance(API_KEYS_V2[0], str):
    try:
        API_KEYS_V2 = json.loads(API_KEYS_V2[0])
    except json.JSONDecodeError:
        raise RuntimeError("API_KEYS_V2 list contains invalid JSON")

if not isinstance(API_KEYS_V2, dict):
    raise RuntimeError("API_KEYS_V2 must be a JSON object (dict)")


if VERBOSE:
    print(f"API_KEYS_V1:{API_KEYS_V1}:")
    print(f"API_KEYS_V2:{API_KEYS_V2}:")


hostname = socket.gethostname()
if hostname == "frhb96148ds":
    DB_CONFIG = DB_CONFIG_ikoula3
elif hostname == "dhoop-NS5x-NS7xAU":
    DB_CONFIG = DB_CONFIG_ekima

#print(DB_CONFIG)
# Mapping des champs API vers les champs de la base de données


# === CHECK SIREN===

def check_siren_in_db(siren, conn):
    cursor = conn.cursor()
    
    # On compte combien de fois le SIREN apparaît
    cursor.execute(f"SELECT 1 FROM {TABLE_FAST} WHERE siren = %s LIMIT 1", (siren,))
    result = cursor.fetchone()
    
    return result is not None # Retourne True si trouvé, sinon False


# === CHARGEMENT DU RÉFÉRENTIEL INSEE ===
def load_insee_referentiel(csv_path):
    """
    Charge le référentiel INSEE depuis un fichier CSV.
    
    Args:
        csv_path: Chemin vers le fichier CSV
        
    Returns:
        DataFrame avec les colonnes : nom_commune_complet, code_postal, 
        nom_region, code_departement
    """
    df_insee = pd.read_csv(
        csv_path,
        dtype={
            'code_postal': str,  # Important : garder en string pour éviter la perte des 0 initiaux
            'code_departement': str
        },
        encoding='utf-8'
    )
    
    # Nettoyer les espaces
    df_insee['nom_commune_complet'] = df_insee['nom_commune_complet'].str.strip()
    df_insee['nom_region'] = df_insee['nom_region'].str.strip()
    df_insee['code_postal'] = df_insee['code_postal'].str.strip()
    df_insee['code_departement'] = df_insee['code_departement'].str.strip()
    
    return df_insee



# Chargement du référentiel INSEE (à faire une seule fois au démarrage)
df_insee = load_insee_referentiel('communes-departement-region.csv')

def insert_api_log(timestamp, request_json, duration, response_json, ip_address):
    """
    Inserts an API call log into the LogAPI_bot table
    """
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        query = """
        INSERT INTO LogAPI_bot (timestamp, request_json, duration, response_json, ip_address)
        VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(query, (
            timestamp,
            json.dumps(request_json, ensure_ascii=False),
            duration,
            json.dumps(response_json, ensure_ascii=False), 
            ip_address
        ))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        logger.error("Failed to log API call: %s", e)


def filter_location_by_hierarchy(location_data, df_insee):
    # Normaliser les inputs en listes
    regions = location_data.get('region') or []
    regions = regions if isinstance(regions, list) else [regions]
    
    departements = location_data.get('departement') or []
    departements = departements if isinstance(departements, list) else [departements]
    
    post_codes = location_data.get('post_code') or []
    post_codes = post_codes if isinstance(post_codes, list) else [post_codes]
    
    cities = location_data.get('city') or []
    cities = cities if isinstance(cities, list) else [cities]

    # --- HIERARCHY LOGIC ---
    # Rule: If a child is inside a parent, drop the parent (keep the precision).
    # If they are independent, keep both (unless your rule is to strictly keep only one level)

    # 1. Check Cities vs Others
    if cities:
        df_cities = df_insee[df_insee['nom_commune_complet'].isin(cities)]
        if not df_cities.empty:
            # Get parents of these cities
            c_depts = df_cities['code_departement'].unique().tolist()
            c_regions = df_cities['nom_region'].unique().tolist()
            c_cps = df_cities['code_postal'].unique().tolist()

            # Remove parents if they contain these cities
            departements = [d for d in departements if d not in c_depts]
            regions = [r for r in regions if r not in c_regions]
            post_codes = [cp for cp in post_codes if cp not in c_cps]

    # 2. Check Postcodes vs Others
    if post_codes:
        df_cp = df_insee[df_insee['code_postal'].isin(post_codes)]
        if not df_cp.empty:
            cp_depts = df_cp['code_departement'].unique().tolist()
            cp_regions = df_cp['nom_region'].unique().tolist()

            departements = [d for d in departements if d not in cp_depts]
            regions = [r for r in regions if r not in cp_regions]

    # 3. Check Departements vs Regions
    if departements and regions:
        df_dept = df_insee[df_insee['code_departement'].isin(departements)]
        if not df_dept.empty:
            d_regions = df_dept['nom_region'].unique().tolist()
            regions = [r for r in regions if r not in d_regions]

    return {
        'region': regions if regions else None,
        'departement': departements if departements else None,
        'post_code': post_codes if post_codes else None,
        'city': cities if cities else None
    }

def get_db_connection():
    """Établit une connexion à la base de données MySQL"""
    try:
        conn = mysql.connector.connect(**DB_CONFIG)
        return conn
    except Error as e:
        logger.exception(f"Erreur de connexion à Mysql")

        print(f"Erreur de connexion à MySQL: {e}")
        raise


def strip_activite_condition(sql: str) -> str:
    #print(f"sql before strip: {sql}")
    
    # On remplace la condition spécifique par une condition toujours vraie
    # Cela évite les problèmes de syntaxe (AND AND, WHERE AND, etc.)
    sql = re.sub(
        r"Activite_entreprise\s+IN\s*\([^)]*\)",
        "1=1",
        sql,
        flags=re.IGNORECASE
    )

    # Nettoyage des espaces doubles pour faire propre
    sql = re.sub(r"\s+", " ", sql).strip()

    #print(f"retour sql stripped: {sql}")
    return sql


def format_sql_for_debug(query: str, params: List[Any]) -> str:
    """
    Returns a printable SQL query with parameters interpolated.
    ⚠️ FOR DEBUGGING ONLY – NEVER execute this string.
    """

    def format_value(value: Any) -> str:
        if value is None:
            return "NULL"

        if isinstance(value, bool):
            return "1" if value else "0"

        if isinstance(value, (int, float)):
            return str(value)

        if isinstance(value, (date, datetime)):
            return f"'{value.isoformat()}'"

        # ✅ LIST → SQL IN (...)
        if isinstance(value, (list, tuple)):
            formatted_items = []
            for v in value:
                if v is None:
                    formatted_items.append("NULL")
                else:
                    formatted_items.append(
                        "'" + str(v).replace("'", "''") + "'"
                    )
            return "(" + ",".join(formatted_items) + ")"

        # STRING
        return "'" + str(value).replace("'", "''") + "'"

    formatted_query = query
    for param in params:
        formatted_query = formatted_query.replace("?", format_value(param), 1)

    return formatted_query


def add_scalar_or_list_filter(where_clauses, params, field_name, value):
    """
    Adds SQL filter for a scalar or list value.
    - scalar -> field = ?
    - list    -> field IN (?, ?, ...)
    """
    if value is None:
        return

    if isinstance(value, list):
        if not value:
            return
        placeholders = ",".join(["?"] * len(value))
        where_clauses.append(f"{field_name} IN ({placeholders})")
        params.extend(value)
    else:
        where_clauses.append(f"{field_name} = ?")
        params.append(value)

from typing import Dict, Any, List

def build_query_legal(criteria: Dict[str, Any], flag_count = False) -> tuple[str, List[Any]]:
    """
    Builds the SQL query and parameters from targeting criteria.
    """
    where_clauses: List[str] = []
    params: List[Any] = []

    # ==============================
    # 1. Table Choice (Headquarters)
    # ==============================
    table = TABLE_ALL
    legal_criteria = criteria.get('legal_criteria', {})
    if legal_criteria.get('present') and legal_criteria.get('headquarters') is True:
        table = TABLE_FAST

    # ==============================
    # 2. Location (Hierarchical + OR logic)
    # ==============================
    if criteria.get('location', {}).get('present'):
        loc = criteria['location']
        filtered_loc = filter_location_by_hierarchy(loc, df_insee)
        
        loc_clauses = []
        # Process each level; if data exists, add to the OR group
        for key in ['city', 'post_code', 'departement', 'region']:
            values = filtered_loc.get(key)
            if values:
                if not isinstance(values, list):
                    values = [values]
                
                placeholders = ", ".join(["?"] * len(values))
                field_name = FIELD_MAPPING[key]
                loc_clauses.append(f"{field_name} IN ({placeholders})")
                params.extend(values)

        # Wrap all location filters in (OR) to avoid breaking other AND filters
        if loc_clauses:
            where_clauses.append(f"({' OR '.join(loc_clauses)})")

    # ==============================
    # 3. Activity
    # ==============================
    if criteria.get('activity', {}).get('present'):
        act = criteria['activity']
        if act.get('activity_codes_list'):
            codes = act['activity_codes_list']
            placeholders = ','.join(['?'] * len(codes))
            where_clauses.append(f"{FIELD_MAPPING['activity_codes_list']} IN ({placeholders})")
            params.extend(codes)

    # ==============================
    # 4. Company Size
    # ==============================
    if criteria.get('company_size', {}).get('present'):
        size = criteria['company_size']
        value = size.get('employees_number_range')
        if value:
            if isinstance(value, list):
                placeholders = ",".join(["?"] * len(value))
                where_clauses.append(f"{FIELD_MAPPING['employees_number_range']} IN ({placeholders})")
                params.extend(value)
            else:
                where_clauses.append(f"{FIELD_MAPPING['employees_number_range']} = ?")
                params.append(value)

    # ==============================
    # 5. Financial Criteria
    # ==============================
    if criteria.get('financial_criteria', {}).get('present'):
        fin = criteria['financial_criteria']
        for field in ['turnover', 'net_profit', 'profitability']:
            if fin.get(field) is not None:
                if fin.get(f'{field}_sup', True):
                    where_clauses.append(f"{FIELD_MAPPING[field]} >= ?")
                    params.append(fin[field])
                if fin.get(f'{field}_inf', False):
                    where_clauses.append(f"{FIELD_MAPPING[field]} <= ?")
                    params.append(fin[field])

    # ==============================
    # 6. Legal Criteria
    # ==============================
    if legal_criteria.get('present'):
        if legal_criteria.get('legal_category'):
            where_clauses.append(f"{FIELD_MAPPING['legal_category']} = ?")
            params.append(legal_criteria['legal_category'])

        # Creation Date
        if legal_criteria.get('company_creation_date_threshold'):
            d_field = FIELD_MAPPING['company_creation_date_threshold']
            val = legal_criteria['company_creation_date_threshold']
            if legal_criteria.get('company_creation_date_sup'):
                where_clauses.append(f"{d_field} >= ?")
                params.append(val)
            if legal_criteria.get('company_creation_date_inf'):
                where_clauses.append(f"{d_field} <= ?")
                params.append(val)

        # Capital
        if legal_criteria.get('capital') is not None:
            c_field = FIELD_MAPPING['capital']
            val = legal_criteria['capital']
            if legal_criteria.get('capital_threshold_sup'):
                where_clauses.append(f"{c_field} >= ?")
                params.append(val)
            if legal_criteria.get('capital_threshold_inf'):
                where_clauses.append(f"{c_field} <= ?")
                params.append(val)

        if legal_criteria.get('subsidiaries_number') is not None:
            where_clauses.append(f"{FIELD_MAPPING['subsidiaries_number']} >= ?")
            params.append(legal_criteria['subsidiaries_number'])

    # ==============================
    # Final Query Assembly
    # ==============================
    if flag_count:
        query = f"SELECT COUNT(*) AS count FROM {table}"
    else:
        query = f"SELECT siren FROM {table}"

    if where_clauses:
        query += " WHERE " + " AND ".join(where_clauses)

    return query, params




def check_sql_params(query: str, params: List) -> None:
    """
    Vérifie que le nombre de paramètres correspond au nombre de placeholders '?' dans la requête SQL.
    Lève une exception si un mismatch est détecté.
    
    Args:
        query (str): La requête SQL avec placeholders '?'
        params (List): La liste des paramètres associés
    
    Raises:
        ValueError: Si le nombre de placeholders et de paramètres ne correspond pas
    """
    num_placeholders = query.count("?")
    num_params = len(params)
    
    if num_placeholders != num_params:
        raise ValueError(
            f"SQL placeholder mismatch detected!\n"
            f"Placeholders: {num_placeholders}, Params: {num_params}\n"
            f"Query: {query}\n"
            f"Params: {params}"
        )


def convert_employees_range_to_salaries(criteria: dict) -> dict:
    """
    Convertit 'X to Y employees' → 'X a Y salaries'
    Supporte string OU liste de strings
    """
    company_size = criteria.get("company_size", {})

    if not company_size.get("present"):
        return criteria

    value = company_size.get("employees_number_range")
    if not value:
        return criteria

    def convert(text: str) -> str:
        text = re.sub(r'\bto\b', 'a', text, flags=re.IGNORECASE)
        text = re.sub(r'\bemployees\b', 'salaries', text, flags=re.IGNORECASE)
        text = re.sub(r'\bemployee\b', 'salarie', text, flags=re.IGNORECASE)

        return text

    # ✅ Case 1: list of ranges
    if isinstance(value, list):
        company_size["employees_number_range"] = [
            convert(v) for v in value if isinstance(v, str)
        ]

    # ✅ Case 2: single string
    elif isinstance(value, str):
        company_size["employees_number_range"] = convert(value)

    return criteria

def test_criteria_mismatches(criteria: Dict) -> List[str]:
    """
    Analyse le JSON des critères et détecte les combinaisons impossibles
    qui pourraient générer un mismatch de paramètres SQL.
    
    Returns:
        List[str]: Liste de messages d'erreur sur les critères incohérents.
    """
    errors = []

    # Check legal criteria
    legal = criteria.get('legal_criteria', {})
    if legal.get('present'):
        # company_creation_date
        if legal.get('company_creation_date_threshold'):
            sup = legal.get('company_creation_date_sup', False)
            inf = legal.get('company_creation_date_inf', False)
            if sup and inf:
                errors.append(
                    "company_creation_date_threshold: both 'sup' and 'inf' are True — check if intended."
                )
            if (sup or inf) and not legal.get('company_creation_date_threshold'):
                errors.append(
                    "company_creation_date_threshold is missing but 'sup' or 'inf' is True"
                )

        # capital thresholds
        if legal.get('capital') is not None:
            sup = legal.get('capital_threshold_sup', False)
            inf = legal.get('capital_threshold_inf', False)
            if (sup or inf) and legal.get('capital') is None:
                errors.append(
                    "capital_threshold_sup or _inf is True but 'capital' is None"
                )
        else:
            if legal.get('capital_threshold_sup') or legal.get('capital_threshold_inf'):
                errors.append(
                    "'capital_threshold_sup' or '_inf' is True but 'capital' is missing"
                )
        
        # Add other dual-flag checks here as needed
        # e.g., financial criteria sup/inf, etc.

    # You can extend to other sections like company_size, financial_criteria if needed
    
    return errors



def lemmatize_expression(expression: str, conn) -> str:
    cursor = conn.cursor(dictionary=True)

    """
    Lemmatize each word in the expression using the DicoFrance table.
    Returns the lemmatized expression as a string.
    """
    if not expression:
        return ""

    # Split words by space
    words = expression.split()

    # Lowercase for consistency
    words_lower = [w.lower() for w in words]

    # Batch query to fetch lemmas
    format_strings = ",".join(["%s"] * len(words_lower))
    sql = f"SELECT entree, lemme FROM DicoFrance WHERE entree IN ({format_strings})"
    cursor.execute(sql, tuple(words_lower))
    results = cursor.fetchall()

    # Map: word -> lemma
    lemma_map = {r['entree']: r['lemme'] for r in results if r['lemme']}

    # Replace words with lemma if found
    lemmatized_words = [lemma_map.get(w, w) for w in words_lower]

    # Join back into a string
    return " ".join(lemmatized_words)

def remove_stop_words_french(texte, supprime_accent, verbose=False):
    """
    Remove French stop words and words shorter than MIN_FULLTEXT_LENGTH.
    """
    if texte is None:
        return "#"

    words = texte.split()
    result = []

    for w in words:
        original_w = w
        if supprime_accent:
            w, _ = removeaccent(w, verbose)
        w_lower = w.lower()

        # Skip stopwords and short words
        if w_lower not in TAB_STOPWORDS and len(w_lower) >= MIN_FULLTEXT_LENGTH:
            result.append(original_w)
        else:
            if verbose:
                print(f"Skipped: {w_lower}")

    return " ".join(result)


def normalize_french_text(text: str) -> tuple[str, str]:
    if not text:
        return "", ""

    # Lowercase
    text = text.lower()

    # Normalize accents (é → e)
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")

    # Remove French elisions: d', l', qu', etc.
    text = re.sub(FRENCH_ELISIONS, "", text)

    # Remove remaining apostrophes
    text = re.sub(r"[’'`]", " ", text)

    # Keep dash (hyphen)
    text = re.sub(r"[^a-z0-9\s\-]", " ", text)

    # Normalize multiple dashes
    text = re.sub(r"-{2,}", "-", text)

    # Normalize spaces around dashes
    text = re.sub(r"\s*-\s*", "-", text)

    # Collapse spaces
    text = re.sub(r"\s+", " ", text).strip()

    # Variant 1: keep dash (for LIKE / exact)
    text_with_dash = text

    # Variant 2: remove dash (for FULLTEXT)
    text_without_dash = text.replace("-", " ")

    return text_with_dash, text_without_dash

def removeaccent(word, verbose=False):
    normalized = unicodedata.normalize("NFD", word)
    no_accent = "".join(c for c in normalized if unicodedata.category(c) != "Mn")
    if verbose:
        print(f"{word} -> {no_accent}")
    return no_accent, word



def count_semantic(original_request, debug_sql, conn, flag_count=False):

    # print(f"original_request:{original_request}")
    #print(f"debug_sql debut fct count_semantic:{debug_sql}")
    idx = debug_sql.upper().rfind("WHERE")
    if idx == -1:
        where_cmd_with_and      = "AND 1 = 1"
        where_cmd_without_and   = "1 = 1"
    else: 
        where_cmd_with_and = "AND " + debug_sql[idx + len("WHERE"):].strip()
        where_cmd_without_and = debug_sql[idx + len("WHERE"):].strip()

    #print(f"****where_cmd_with_and:{where_cmd_with_and}")
    # ---- 1️⃣ Your lemmatization function ----
    def lemmatize(text: str) -> str:
        # Replace with your actual lemmatization
        return text.lower()


    # ---- 3️⃣ Lemmatize the expression ----
    # print(f"original_request:{original_request}")
    # print("normalisze")
    # print(normalize_french_text(original_request), True, False)
    original_request_normalized_with_dash,original_request_normalized_without_dash = normalize_french_text(original_request)

    if "-" in original_request:
        expr_lem_with_dash      = lemmatize_expression( remove_stop_words_french(original_request_normalized_with_dash, True, False), conn)  # e.g., "directeur immobilier"
        expr_lem_without_dash   = lemmatize_expression( remove_stop_words_french(original_request_normalized_without_dash, True, False), conn)  # e.g., "directeur immobilier"
        # print(f"expr_lem_with_dash:{expr_lem_with_dash}")
        # print(f"expr_lem_without_dash:{expr_lem_without_dash}")
    else:
        expr_lem = lemmatize_expression( remove_stop_words_french(original_request_normalized_without_dash, True, False), conn)  # e.g., "directeur immobilier"
        expr_boolean = ' '.join(f'+{word}' for word in expr_lem.split())
        #print(f"expr_lem:{expr_lem}")
   

    # ---- 4️⃣ Convert to Boolean mode format ----
    # Example result: "+directeur +immobilier"

    cursor = conn.cursor()

   
    # ---- 6️⃣ SQL query with HAVING ----
    if "-" in original_request:
         query = f"""
                SELECT DISTINCT siren AS matching_companies
                FROM (

                /* -------------------------------
                   1️⃣ AFNIC
                -------------------------------- */
                SELECT
                    s.siren
                FROM (
                    SELECT
                        COALESCE(
                            sirentrouve,
                            sirentrouveaadresse,
                            sirentrouve_semantique
                        ) AS siren
                    FROM {TABLE_AFNIC}
                    WHERE MATCH(
                            title_lemmatise,
                            description_lemmatise,
                            keywords_lemmatise,
                            TexteHome_lemmatise,
                            keyword_nomdedomaine_lemmatise
                          )
                        AGAINST('{expr_lem_without_dash}' IN BOOLEAN MODE)
                      AND (
                            (sirentrouve IS NOT NULL and sirentrouve != 0)
                            OR (sirentrouveaadresse IS NOT NULL and sirentrouveaadresse != 0)
                            OR (sirentrouve_semantique IS NOT NULL and sirentrouve_semantique != 0)
                          )
                           AND (
                            title LIKE '%{expr_lem_with_dash}%'
                            OR description LIKE '%{expr_lem_with_dash}%'
                            OR keywords LIKE '%{expr_lem_with_dash}%'
                            OR TexteHome LIKE '%{expr_lem_with_dash}%'
                            OR keywords LIKE '%{expr_lem_with_dash}%'
                            )
                ) af
                INNER JOIN {TABLE_FAST} s
                    ON s.siren = af.siren
                WHERE
                    {where_cmd_without_and}

                UNION ALL

                /* -------------------------------
                   2️⃣ BODACC
                -------------------------------- */
                SELECT
                    s.siren
                FROM Bodacc_Light{MOIS_ANNEE}_full b
                INNER JOIN {TABLE_FAST} s
                    ON s.siren = b.siren
                WHERE MATCH(b.Objet_Social_lemmatisee)
                      AGAINST('{expr_lem_without_dash}' IN BOOLEAN MODE)
                  {where_cmd_with_and}
                AND Objet_Social LIKE '%{expr_lem_with_dash}%'


                UNION ALL

                /* -------------------------------
                   3️⃣ SIRENE
                -------------------------------- */
                SELECT
                    s.siren
                FROM {TABLE_FAST} s
                WHERE MATCH(s.Nom_entreprise_lemmatise, s.Nom_enseigne)
                      AGAINST('{expr_lem_without_dash}' IN BOOLEAN MODE)
                  {where_cmd_with_and}
                   AND (
                    Nom_entreprise LIKE '{expr_lem_with_dash}'
                 OR Nom_enseigne LIKE '{expr_lem_with_dash}'
              )

            ) t
            WHERE siren IS NOT NULL and siren != 0
                """
    else:

        query = f"""
        SELECT DISTINCT siren AS matching_companies
        FROM (

        /* -------------------------------
           1️⃣ AFNIC
        -------------------------------- */
        SELECT
            s.siren
        FROM (
            SELECT
                COALESCE(
                    sirentrouve,
                    sirentrouveaadresse,
                    sirentrouve_semantique
                ) AS siren
            FROM {TABLE_AFNIC}
            WHERE MATCH(
                    title_lemmatise,
                    description_lemmatise,
                    keywords_lemmatise,
                    TexteHome_lemmatise,
                    keyword_nomdedomaine_lemmatise
                  )
                AGAINST('{expr_boolean}' IN BOOLEAN MODE)
              AND (
                    (sirentrouve IS NOT NULL and sirentrouve != 0)
                    OR (sirentrouveaadresse IS NOT NULL and sirentrouveaadresse != 0)
                    OR (sirentrouve_semantique IS NOT NULL and sirentrouve_semantique != 0)
                  )
        ) af
        INNER JOIN {TABLE_FAST} s
            ON s.siren = af.siren
        WHERE
            {where_cmd_without_and}

        UNION ALL

        /* -------------------------------
           2️⃣ BODACC
        -------------------------------- */
        SELECT
            s.siren
        FROM Bodacc_Light{MOIS_ANNEE}_full b
        INNER JOIN {TABLE_FAST} s
            ON s.siren = b.siren
        WHERE MATCH(b.Objet_Social_lemmatisee)
              AGAINST('{expr_boolean}' IN BOOLEAN MODE)
          {where_cmd_with_and}

        UNION ALL

        /* -------------------------------
           3️⃣ SIRENE
        -------------------------------- */
        SELECT
            s.siren
        FROM {TABLE_FAST} s
        WHERE MATCH(s.Nom_entreprise_lemmatise, s.Nom_enseigne)
              AGAINST('{expr_boolean}' IN BOOLEAN MODE)
          {where_cmd_with_and}

    ) t
    WHERE siren IS NOT NULL and siren != 0
        """

    # ---- 7️⃣ Print query for debugging ----
    if VERBOSE:
        print("DEBUG: SQL Query to execute:")
        print(query)

    # ---- 8️⃣ Execute query ----
    cursor.execute(query)
    result = cursor.fetchall()
    cleaned_list = [siren[0] for siren in result]
    # print(f"cleaned_list: {cleaned_list}")

    matching_rows = len(cleaned_list)
    # print(f"matching_rows: {matching_rows}")


    #print(f"Number of matching rows: {matching_rows}")
    return matching_rows, cleaned_list


def require_api_key_v1(func):
    @wraps(func)
    def decorated(*args, **kwargs):
        # Normalize headers to lowercase for case-insensitive lookup
        headers = {k.lower(): v for k, v in request.headers.items()}

        # Try X-Api-Key first
        api_key = headers.get("x-api-key")

        # Fallback to Authorization header: Bearer <token>
        if not api_key:
            auth_header = headers.get("authorization")
            if auth_header and auth_header.lower().startswith("bearer "):
                api_key = auth_header.split(" ", 1)[1]

        if VERBOSE:

            print("All headers received:", dict(request.headers))
            print(f"API key received:{api_key}:")
            print(f"API_KEYS_V1:{API_KEYS_V1}:")
            print(not api_key or api_key not in API_KEYS_V1)
            print(api_key not in API_KEYS_V1)
        if not api_key or api_key not in API_KEYS_V1:
            return jsonify({"error": "Unauthorized: Invalid or missing API key"}), 401

        return func(*args, **kwargs)
    return decorated




def require_api_key_v2(func):
   
    AUTH_WINDOW = 300  # 5 minutes

    @wraps(func)
    def decorated(*args, **kwargs):
        headers = {k.lower(): v for k, v in request.headers.items()}

        api_key     = headers.get("x-api-key")
        timestamp   = headers.get("x-timestamp")
        signature   = headers.get("x-signature")

        if not api_key or not timestamp or not signature:
            return jsonify({"error": "Missing authentication headers"}), 401

        #print(os.getenv("API_KEYS_V2", "{}"))
        API_KEYS = json.loads(os.getenv("API_KEYS_V2", "{}"))

        secret = API_KEYS.get(api_key)
        if not secret:
            return jsonify({"error": "Invalid API key"}), 401

        # Timestamp validation
        try:
            timestamp = int(timestamp)
        except ValueError:
            return jsonify({"error": "Invalid timestamp"}), 401

        if abs(time.time() - timestamp) > AUTH_WINDOW:
            return jsonify({"error": "Expired request"}), 401

        # Signature validation
        message = f"{api_key}{timestamp}".encode()
        expected = hmac.new(
            secret.encode(),
            message,
            hashlib.sha256
        ).hexdigest()

        if not hmac.compare_digest(expected, signature):
            return jsonify({"error": "Invalid signature"}), 401

        return func(*args, **kwargs)

    return decorated

def count_companies_logic(criteria: dict):
    """
    Pure business logic.
    Returns a Python dict ONLY.
    """
    start_time = time.perf_counter()
    timestamp = datetime.now()

    response = None

    try:
        if not criteria:
            raise ValueError("Request body is empty")

        conn = get_db_connection()
        cursor = conn.cursor(dictionary=True)

        criteria_dict = criteria.get("criteria", criteria)
        security_block = criteria_dict.get("security", {})
        if security_block:
            ip_address = security_block.get("ip_address")
        else:
            ip_address = "0.0.0.0"

        criteria_dict = convert_employees_range_to_salaries(criteria_dict)
        # Step 1: Get the inner dictionary (execution_mode)
        execution_block = criteria_dict.get("execution_mode", {})

        # Step 2: Get the specific value (output_type)
        # We provide "count" as a second argument so it has a default value
        mode = execution_block.get("output_type", "count")
        original_activity_request = (
            criteria_dict.get('activity', {})
            .get('original_activity_request', [])
        )
        semantic_count_requested = (
            criteria_dict.get('activity', {})
            .get('semantic_count_requested', False)
        )
        #print(f"semantic_count_requested:{semantic_count_requested}")
        # --- Total count legal ---

        # doc the function includes the swith count/list of siren
        query, params = build_query_legal(criteria_dict, mode == "count")
        check_sql_params(query, params)

        debug_sql = format_sql_for_debug(query, params)

        if VERBOSE:
            print(f"debug_sql:{debug_sql}")
        cursor.execute(debug_sql)


        #print(f"mode:{mode}")

        if mode == "count":

            result = cursor.fetchone()

            total_count_legal = result['count'] if result else 0
       
            list_siren_semantic = None
            # --- Semantic ---

            # doc on laisse le double run le temps d'adaptation du front

            if (semantic_count_requested or original_activity_request):
                total_count_semantic, list_siren_semantic = count_semantic(original_activity_request, strip_activite_condition(debug_sql),conn, True)
            else:
                total_count_semantic = 0

            # --- Individual counts ---
            activity_individual_counts = {}
            activity_codes = criteria_dict.get('activity', {}).get('activity_codes_list', [])

            for code in activity_codes:
                criteria_single = copy.deepcopy(criteria_dict)
                criteria_single['activity'] = {
                    'present': True,
                    'activity_codes_list': [code]
                }

                # doc the function includes the swith count/list of siren
                query_code, params_code = build_query_legal(criteria_single, True)
                debug_sql_code          = format_sql_for_debug(query_code, params_code)

                cursor.execute(debug_sql_code)
                result_code = cursor.fetchone()

                activity_individual_counts[code] = {
                    "count_legal": result_code['count'] if result_code else 0
                }


            response = {
                'count_legal': total_count_legal,
                'count_semantic': total_count_semantic,
                'activity_individual_counts': activity_individual_counts or None,
                'debug_sql': debug_sql
            }

        elif mode == "display":

            #print("mode : display")
            # doc without count siren list should be provided in the sql command
            result = cursor.fetchall()
            #print(f"toto result:{result}")
            if isinstance(result, list):
                siren_list_legal = [
                    row['siren'] if isinstance(row, dict) else row 
                    for row in result
                ]
            else:
                siren_list_legal = []

            # Votre logique de flag
            #print(f"semantic_count_requested:{semantic_count_requested}")
            list_siren_semantic = None
            # doc on laisse le double run le temps d'adaptation du front
            if (semantic_count_requested or original_activity_request):
                # doc calcul des siren issus de la recherche sémantique.
                total_count_semantic, list_siren_semantic = count_semantic(original_activity_request, strip_activite_condition(debug_sql),conn, False)
            else:
                total_count_semantic = 0
                list_siren_semantic = siren_list_legal
            
            #print(f"list_siren_semantic:{list_siren_semantic}")
            response = {
                "count": len(siren_list_legal),
                "results_semantic": list_siren_semantic,
                "results_legal": siren_list_legal
            }

        elif mode == "big_file":
            result = cursor.fetchall()
            #print(f"result:{result}")
            siren_list = [row['siren'] for row in result]

            # Mapping to a dictionary first
            response = {
                "count": len(siren_list),
                "siren_list": siren_list
            }

        return response

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

        duration = time.perf_counter() - start_time
        try:
            #print(f"ip_address:{ip_address}")
            insert_api_log(timestamp, criteria, duration, response, ip_address)
        except Exception:
            logger.exception(f"Erreur insertion API log")

            pass


def get_company_info(list_siren, origin_label):

    conn = get_db_connection()
    cursor = conn.cursor(dictionary=True)  

    try:
        # 1. Sécurité : vérifier si la liste est vide
        if not list_siren:
            #print("return")
            # Retourne un dictionnaire cohérent avec la structure attendue
            return {"data": [], "metadata": {"total_found": 0, "requested_count": 0, "price": 0}}

        # 2. Générer les placeholders
        placeholders = ', '.join(['%s'] * len(list_siren))
        
        sql = f"""
            SELECT 
                siren, Nom_entreprise, Commune, Code_postal, Departement, Region, 
                Activite_entreprise, Tranche_effectif_entreprise, 
                Date_creation_entreprise, Capital, CA_le_plus_recent, 
                Resultat_net_le_plus_recent, Rentabilite_la_plus_recente
            FROM {TABLE_ALL}
            WHERE siren IN ({placeholders}) AND Siege_entreprise = 'oui' 
            LIMIT {LIMIT_DISPLAY_INFO}
        """

        cursor.execute(sql, tuple(list_siren)) 
        rows = cursor.fetchall()

        # 4. Data Transformation
        rows_display = rows[:LIMIT_DISPLAY_INFO]
        price = len(rows) * UNITARY_PRICE_LEGAL_INFOS
       # --- Dans votre fonction get_company_info ---
        rows_display = []
        for row in rows:
            # On définit l'ordre manuellement ici
            new_row = {
                "Origine de la recherche": origin_label,
                "siren": row.get("siren"),
                "Nom_entreprise": row.get("Nom_entreprise"),
                "Commune": row.get("Commune"),
                "Code_postal": row.get("Code_postal"),
                "Departement": row.get("Departement"),
                "Region": row.get("Region"),
                "Activite_entreprise": row.get("Activite_entreprise"),
                "Tranche_effectif_entreprise": row.get("Tranche_effectif_entreprise"),
                "Date_creation_entreprise": row.get("Date_creation_entreprise"),
                "Capital": row.get("Capital"),
                "CA_le_plus_recent": row.get("CA_le_plus_recent"),
                "Resultat_net_le_plus_recent": row.get("Resultat_net_le_plus_recent"),
                "Rentabilite_la_plus_recente": row.get("Rentabilite_la_plus_recente")
            }
            rows_display.append(new_row)
        

        output = {
            "metadata": {
                "total_found": len(rows),
                "requested_count": len(list_siren),
                "price": math.ceil(price)
            },
            "data": rows_display,
        }
        return output # On renvoie l'objet dict proprement

    finally:
        cursor.close()
        conn.close()


def get_company_file(file_link, file_path, criteria, list_siren, stripe_id):
    #print(f"Critéres reçus : {criteria}")
    
    
    # On récupère les champs dans execution_mode

    exec_mode = criteria.get('execution_mode', {})
    #raw_email = exec_mode.get('email_address', 'inconnu')

    #print(f"exec_mode:{exec_mode}")
    web_email_req = exec_mode.get('web_site_email_requested', False)
    web_phone_req = exec_mode.get('web_site_phone_requested', False)
    # print(f"web_email_req:{web_email_req}")
    # print(f"web_phone_req:{web_phone_req}")

   

    # Initialisation pour éviter les UnboundLocalError
    output = {"metadata": {"total_found": 0, "file_link": None}}
    # Assurez-vous que ce dossier existe sur votre serveur
    #print(f"file_name:{file_name}")

    conn = None
    cursor = None

    try:
        conn = get_db_connection()
        # dictionary=True est crucial pour que Pandas transforme les lignes en colonnes
        cursor = conn.cursor(dictionary=True)
        
        # 1. Extraction de la liste de SIREN
        if isinstance(list_siren, dict):
            list_siren = list_siren.get('siren_list', [])

        if not list_siren:
            return json.dumps({"metadata": {"total_found": 0, "file_link": None}})

        placeholders = ', '.join(['%s'] * len(list_siren))

        if (web_email_req and not web_phone_req):
            #print("web_email_req and not web_phone_req")
            # 2. Requête SQL avec email
            sql = f"""
                SELECT 
                    siren, Commune, Code_postal, Departement, Region, 
                    Activite_entreprise, Libelle_activite_entreprises, Tranche_effectif_entreprise, 
                    Date_creation_entreprise, Capital, CA_le_plus_recent, 
                    Resultat_net_le_plus_recent, Rentabilite_la_plus_recente, Best_email
                FROM {TABLE_ALL}
                WHERE siren IN ({placeholders}) AND Siege_entreprise = 'oui'  AND Best_Email like '%@%'
            """
        elif (web_phone_req and not web_email_req):
            #print("web_phone_req and not web_email_req")

            # 2. Requête SQL avec telephone
            sql = f"""
                SELECT 
                    siren, Commune, Code_postal, Departement, Region, 
                    Activite_entreprise, Libelle_activite_entreprises, Tranche_effectif_entreprise, 
                    Date_creation_entreprise, Capital, CA_le_plus_recent, 
                    Resultat_net_le_plus_recent, Rentabilite_la_plus_recente, Telephone_fixe, Telephone_mobile
                FROM {TABLE_ALL}
                WHERE siren IN ({placeholders}) AND Siege_entreprise = 'oui' AND Presence_numeros_de_telephone = 'oui'
            """

        elif (web_email_req and web_phone_req):
            # 2. Requête SQL avec telephone
            #print("web_email_req and web_phone_req")

            sql = f"""
                SELECT 
                    siren, Commune, Code_postal, Departement, Region, 
                    Activite_entreprise, Libelle_activite_entreprises, Tranche_effectif_entreprise, 
                    Date_creation_entreprise, Capital, CA_le_plus_recent, 
                    Resultat_net_le_plus_recent, Rentabilite_la_plus_recente, Best_Email, Telephone_fixe, Telephone_mobile
                FROM {TABLE_ALL}
                WHERE siren IN ({placeholders}) AND Siege_entreprise = 'oui' AND Best_Email like '%@%' AND Presence_numeros_de_telephone = 'oui'
            """
        else:
            # 2. Requête SQL
            #print("not web_email_req and not web_phone_req")
            sql = f"""
                SELECT 
                    siren, Commune, Code_postal, Departement, Region, 
                    Activite_entreprise, Libelle_activite_entreprises, Tranche_effectif_entreprise, 
                    Date_creation_entreprise, Capital, CA_le_plus_recent, 
                    Resultat_net_le_plus_recent, Rentabilite_la_plus_recente
                FROM {TABLE_ALL}
                WHERE siren IN ({placeholders}) AND Siege_entreprise = 'oui' 
            """

        if VERBOSE == "yes":
            print(f"sql:{sql}")

        # 3. Exécution
        cursor.execute(sql, tuple(list_siren)) 
        rows = cursor.fetchall()

        if rows:
            df_data = pd.DataFrame(rows)

            # --- PRÉPARATION ONGLET PARAMÉTRAGE ---
            setup_rows = []
            for section, details in criteria.items():
                if isinstance(details, dict):
                    status = "ACTIF" if details.get('present') else "INACTIF"
                    # On filtre 'present' pour ne garder que les valeurs des filtres
                    params = ", ".join([f"{k}: {v}" for k, v in details.items() if k != 'present'])
                    setup_rows.append({
                        "Section": section.upper(), 
                        "Statut": status, 
                        "Détails des filtres": params if params else "Aucun filtre spécifique"
                    })
            df_setup = pd.DataFrame(setup_rows)

            # --- GÉNÉRATION EXCEL ---
            with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                # Création des feuilles
                df_setup.to_excel(writer, index=False, sheet_name='Parametrage')
                df_data.to_excel(writer, index=False, sheet_name='Entreprises')


                # Mise en forme 'Entreprises'
                ws_data = writer.sheets['Entreprises']
                ws_data.freeze_panes = 'A2' # Fige les titres
                
                # Mapping des largeurs personnalisées
                column_widths = {
                    'siren': 15, 'Commune': 30, 'Code_postal': 12, 'Departement': 15,
                    'Region': 30, 'Activite_entreprise': 10, 'Libelle_activite_entreprises': 50,
                    'Tranche_effectif_entreprise': 27, 'Date_creation_entreprise': 20,
                    'Capital': 15, 'CA_le_plus_recent': 20, 'Resultat_net_le_plus_recent': 25,
                    'Rentabilite_la_plus_recente': 30, 'Best_Email': 40, 'Telephone_fixe':40, 'Telephone_mobile':40
                }

                for i, col_name in enumerate(df_data.columns):
                    col_letter = get_column_letter(i + 1) # Utilisation de la fonction importée
                    width = column_widths.get(col_name, 15)
                    ws_data.column_dimensions[col_letter].width = width

                # Mise en forme 'Parametrage'
                ws_setup = writer.sheets['Parametrage']
                ws_setup.column_dimensions['A'].width = 25
                ws_setup.column_dimensions['B'].width = 15
                ws_setup.column_dimensions['C'].width = 80

        else:
            file_link = None

        output = {
            "metadata": {
                "total_found": len(rows)
            }
        }

    except Exception as e:
        if VERBOSE == "yes":
            print(f"Erreur lors de la génération du fichier : {e}")
        output = {"error": True, "message": str(e), "metadata": {"total_found": 0, "file_link": None}}
        logger.error(f"Erreur lors de la génération du fichier : {e}")

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    return json.dumps(output, indent=4, default=str, ensure_ascii=False)



def trafic_control(f):
    @wraps(f) # Important to preserve function metadata
    def decorated_function(*args, **kwargs):
        #print("passage traffic control")
        conn = None
        cursor = None
        try:
            # 1. Setup Database Connection
            conn = get_db_connection()
            cursor = conn.cursor(dictionary=True)

            # 2. Extract IP from incoming request
            data = request.get_json(silent=True) or {}
            security_block = data.get("criteria", data).get("security", {})
            ip_address = security_block.get("ip_address", "0.0.0.0")

            # 3. Query the LogAPI_bot table
            query = """
                SELECT COUNT(*) as count 
                FROM LogAPI_bot 
                WHERE ip_address = %s 
                AND timestamp > NOW() - INTERVAL 1 DAY
            """
            #print(query)
            cursor.execute(query, (ip_address,))
            result = cursor.fetchone()
            #print(f"result:{result}")
            request_count = result['count'] if result else 0

            # 4. Check against threshold
            if request_count >= MAX_DAILY_REQUESTS_NUMBER:
                logger.exception("Il semble que vous ayez entrepris beaucoup de recherches aujourd'hui. Merci de recommencer demain.")

                #print("access denied")
                return jsonify({
                    "status": "denied",
                    "message": "Il semble que vous ayez entrepris beaucoup de recherches aujourd'hui. Merci de recommencer demain."
                }), 429 

        except Exception as e:
            logger.exception(f"Database error: {e}")

            # It's good practice to log the error so you know if the DB fails
            print(f"Database error: {e}")
            # Decide if you want to block or allow if the DB is down
            
        finally:
            # This ensures clean up happens even if the code returns or crashes
            if cursor:
                cursor.close()
            if conn:
                conn.close()

        # 5. If we passed the check, execute the original function
        return f(*args, **kwargs)
        
    return decorated_function

def query_control(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not request.is_json:
            return jsonify({'error': 'Content-Type must be application/json'}), 400

        criteria = request.get_json()

        criteria_dict = criteria.get("criteria", criteria)

        activity = criteria.get("activity", {})
        

        # 1. Extract the original request
        raw_word = activity.get("original_activity_request", "")
        
        # 2. Use your function (unpacking the tuple)
        # We pass VERBOSE (the global you defined earlier) to your function
        clean_word, _ = removeaccent(raw_word, verbose=VERBOSE)
        
        # 3. Final cleaning for comparison
        # We lowercase it because lists of "SANS_ACCENT" are usually lowercase
        final_check_word = clean_word.lower().strip()

        # 4. Check against your parameter list
        if final_check_word in TAB_MOTS_SUP100000_SANS_ACCENT:
            logger.exception(f"{MESSAGE_WORD_TOO_COMMON_1} {final_check_word} {MESSAGE_WORD_TOO_COMMON_2}")
            return jsonify({
                "status": "denied",
                "message": f"{MESSAGE_WORD_TOO_COMMON_1} {final_check_word} {MESSAGE_WORD_TOO_COMMON_2}"
            }), 400

        return f(*args, **kwargs)
    return decorated_function


def memory_guard(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        # Get memory statistics
        memory_info = psutil.virtual_memory()
        current_usage = memory_info.percent  # returns a float like 82.5

        if current_usage > MEMORY_THRESHOLD_PERCENT:
            # Log this internally so you know the server is under stress
            logger.exception(f"CRITICAL: Memory usage at {current_usage}%! Blocking request.")

            if VERBOSE == "yes":
                print(f"CRITICAL: Memory usage at {current_usage}%! Blocking request.")
            
            return jsonify({
                "status": "denied",
                "message": "Le serveur est actuellement très occupé. Merci de réessayer dans quelques instants."
            }), 503  # 503 Service Unavailable is the correct code for a busy server

        return f(*args, **kwargs)
    return decorated_function


def insert_stripe_id_file_link_criteria(new_stripe_id, file_price, billing_post_code, billing_address, billing_city, billing_full_name, local_data_file, data_file_link, company, siren, criteria, conn):

    try:
        cursor = conn.cursor()
        
        # Requête paramétrée pour la sécurité

        # doc Convertir le dict en chaîne JSON
        criteria_json_string = json.dumps(criteria)
        sql = "INSERT INTO billing_stripe SET stripe_id = %s, file_price = %s, criteres_json = %s, billing_post_code = %s, billing_address = %s, billing_city = %s, billing_full_name = %s, company = %s, siren = %s, local_data_file = %s,  data_file_link = %s" 
        cursor.execute(sql, (new_stripe_id, file_price, criteria_json_string, billing_post_code, billing_address, billing_city, billing_full_name, company, siren, local_data_file,  data_file_link))
        
        conn.commit()
        #print(f"✅ insert ID, criteria and file_link: {stripe_id}")
        
    except mysql.connector.Error as err:
        if VERBOSE == "yes":
            print(f"❌ Erreur : {err}")
        logger.exception(f"Erreur insertion données de facturation:{err}")

    finally:
        cursor.close()

def get_invoice_number(stripe_id_value, conn):

    
    purchase_timestamp  = datetime.now()

    try:
        cursor = conn.cursor()
        
        sql = "INSERT INTO invoice_number (purchase_timestamp, stripe_id) VALUES (%s, %s)"
        params = (purchase_timestamp, str(stripe_id_value))
        
       
        cursor.execute(sql, params)
        
        conn.commit()
        
        new_invoice_id = cursor.lastrowid
        if VERBOSE == "yes":
            print(f"✅ Ligne insérée. Facture n° : {new_invoice_id}")

    except mysql.connector.Error as err:
        if VERBOSE == "yes":
            print(f"❌ Erreur lors de l'insertion : {err}")
        conn.rollback() # Annule en cas d'erreur
        logger.exception(f"Erreur génération numero de facture:{err}")

    finally:
        cursor.close()
    return new_invoice_id

def update_payment_invoice(stripe_id, invoice_file_link, conn):

    cursor = None # Initialisation pour éviter l'erreur dans le finally
    try:  
        cursor = conn.cursor()

        query = """
            UPDATE billing_stripe 
            SET invoice_file_link = %s
            WHERE stripe_id = %s
        """
        timestamp = datetime.now()

        # Attention : utilisez bien email_client (le nom de votre argument)
        cursor.execute(query, (invoice_file_link, stripe_id),)
        
        conn.commit()
        
        rows_affected = cursor.rowcount
        return rows_affected > 0
        
    except mysql.connector.Error as err:
        logger.exception(f"Erreur update_payment_invoice:{err}")

        if VERBOSE == "yes":
            print(f"Erreur : {err}")
        return False
    finally:
        # Le finally s'exécute QUOI QU'IL ARRIVE (succès ou erreur)
        if cursor:
            cursor.close()

def update_payment_info_email(stripe_id, email_client, card_owner, conn):
    #print("update_payment_info_email_get_data_files_link")
    cursor = None # Initialisation pour éviter l'erreur dans le finally
    try:  
        cursor = conn.cursor()

        query = """
            UPDATE billing_stripe 
            SET purchase_timestamp = %s, customer_email = %s, card_owner = %s
            WHERE stripe_id = %s
        """
        timestamp = datetime.now()

        # Attention : utilisez bien email_client (le nom de votre argument)
        cursor.execute(query, (timestamp, email_client, card_owner, stripe_id),)
        
        conn.commit()
        
        rows_affected = cursor.rowcount
        return rows_affected > 0
        
    except mysql.connector.Error as err:
        logger.exception(f"Erreur update info payment:{err}")

        if VERBOSE == "yes":
            print(f"Erreur : {err}")
        return False
    finally:
        # Le finally s'exécute QUOI QU'IL ARRIVE (succès ou erreur)
        if cursor:
            cursor.close()

# doc this function check that data file has been generated
def check_data_file_ready(stripe_id):

    #print(f"PATH_LOCAL_DATA_FILE:{PATH_LOCAL_DATA_FILE}")
    directory_to_search = Path(PATH_LOCAL_DATA_FILE)
    search_term         = stripe_id  # Votre identifiant hexa de 20 car.

    if VERBOSE == "yes":
        print(f"Recherche du fichier pour le stripe_id: {search_term}")

    flag_present = False
    # .rglob cherche récursivement
    # On utilise *{search_term}* pour trouver le terme n'importe où dans le nom
    for file_path in directory_to_search.rglob(f"*{search_term}*"):
        if file_path.is_file():
            if VERBOSE == "yes":
                print(f"✅ Fichier trouvé : {file_path.name}")
                print(f"Chemin complet : {file_path.absolute()}")
            flag_present = True
            break  
    if not flag_present:
        logger.error(f"❌ Aucun fichier correspondant trouvé.")
        if VERBOSE == "yes":
            print("❌ Aucun fichier correspondant trouvé.")
    return flag_present
           

def wait_for_data_file(stripe_id, conn):
    """
    Boucle de vérification du fichier pendant 'timeout' secondes.
    """
    start_time = time.time()
    
    if VERBOSE == "yes":
        print(f"⏳ Attente de la production du fichier pour {stripe_id}...")

    while True:
        # 1. On appelle votre fonction de vérification
        if check_data_file_ready(stripe_id):
            if VERBOSE == "yes":
                print("✅ Fichier détecté !")
            return True  # Succès, on sort de la fonction
        
        # 2. Vérification du temps écoulé
        elapsed_time = time.time() - start_time
        if elapsed_time > TIME_OUT_DATA_FILE_PRODUCTION:
            logger.error(f"❌ Erreur : Temps d'attente dépassé ({timeout}s")
            if VERBOSE == "yes":
                print(f"❌ Erreur : Temps d'attente dépassé ({timeout}s)")
            return False  # Timeout atteint, on sort avec un échec
        
        # 3. Pause courte avant la prochaine tentative (pour ne pas saturer le CPU/Disque)
        time.sleep(0.5)

# doc this function get data_file_link info
def get_data_file_link(stripe_id, conn):
    cursor = None
    try:
        cursor = conn.cursor()
        
        # Sélection du champ spécifique
        query = "SELECT local_data_file, data_file_link  FROM billing_stripe WHERE stripe_id = %s"
        cursor.execute(query, (stripe_id,))
        
        result = cursor.fetchone()
        
        # result est un tuple (valeur,), on retourne l'index 0
        if result:
            return result[0], result[1]
        return None, None

    except mysql.connector.Error as err:
        logger.exception(f"❌ Erreur lors de la récupération du lien : {err}")

        if VERBOSE == "yes":
            print(f"Erreur lors de la récupération du lien : {err}")

        return None, None
    finally:
        if cursor:
            cursor.close()



def push_delivery_files(local_data_file, data_remote_file_with_email_path, invoice_local_file, invoice_remote_file_path ):
    if VERBOSE == "yes":
        print(f"🚀 Début du transfert SFTP : {local_path}")
    transport = None
    try:
        # print(f"SFTP_USERNAME:{SFTP_USERNAME}")
        # print(f"SFTP_PASSWORD:{SFTP_PASSWORD}")
        # 1. Établir la connexion de transport
        transport = paramiko.Transport((SFTP_NAME, SFTP_PORT))
        transport.connect(username=SFTP_USERNAME, password=SFTP_PASSWORD)

        # 2. Créer le client SFTP à partir du transport
        sftp = paramiko.SFTPClient.from_transport(transport)

        # 3. Envoyer le fichier data (Local -> Distant)
        # print("*************** totototot")
        # print(f"local_data_file:{local_data_file}")
        # print(f"data_remote_file_with_email_path:{data_remote_file_with_email_path}")
        sftp.put(local_data_file, data_remote_file_with_email_path)
        

        # 4. Envoyer le fichier data (Local -> Distant)
        # print("*************** titititi")
        # print(f"invoice_local_file:{invoice_local_file}")
        # print(f"invoice_remote_file_path:{invoice_remote_file_path}")
        sftp.put(invoice_local_file, invoice_remote_file_path)
        
        if VERBOSE == "yes":
            print(f"✅ Fichier envoyé avec succès vers {remote_path}")
        sftp.close()
        return True

    except Exception as err:
        logger.exception(f"❌ Erreur SFTP: {err}")

        if VERBOSE == "yes":
            print(f"❌ Erreur SFTP : {e}")
        return False
    finally:
        if transport:
            transport.close()
    return status_push_delivery_files


def prepare_logo_for_pdf(logo_path):
    """
    Convertit et optimise une image pour FPDF
    """
    import os
    from PIL import Image
    
    if not os.path.exists(logo_path):
        logger.exception(f"❌ Logo introuvable : {logo_path}")
        return None
    
    try:
        # Ouvrir l'image
        img = Image.open(logo_path)
        #print(f"📸 Image originale : {img.format}, {img.size}, {img.mode}")
        
        # ✅ Convertir RGBA → RGB (enlever la transparence)
        if img.mode == 'RGBA':
            #print("🔄 Conversion RGBA → RGB")
            # Créer un fond blanc
            background = Image.new('RGB', img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3])  # Utiliser le canal alpha comme masque
            img = background
        elif img.mode != 'RGB':
            #print(f"🔄 Conversion {img.mode} → RGB")
            img = img.convert('RGB')
        
        # ✅ Redimensionner (max 2000px de large)
        max_width = 2000
        if img.width > max_width:
            ratio = max_width / img.width
            new_size = (max_width, int(img.height * ratio))
            #print(f"🔄 Redimensionnement : {img.size} → {new_size}")
            img = img.resize(new_size, Image.Resampling.LANCZOS)
        
        # ✅ Sauvegarder en PNG optimisé
        temp_logo = "/tmp/logo_optimized_for_pdf.png"
        img.save(temp_logo, "PNG", optimize=True)
        
        # Vérifier le résultat
        file_size = os.path.getsize(temp_logo)
        if VERBOSE == "yes":
            print(f"✅ Logo optimisé : {temp_logo} ({file_size} octets)")
        
        return temp_logo
        
    except Exception as e:
        logger.exception(f"❌ Erreur lors de la préparation du logo : {e}")
        traceback.print_exc()
        return None


class ProfessionalInvoice(FPDF):
    """Classe personnalisée pour créer des factures professionnelles"""
    
    # Palette de couleurs professionnelle
    COLOR_PRIMARY = (41, 128, 185)      # Bleu professionnel
    COLOR_SECONDARY = (52, 73, 94)      # Gris foncé
    COLOR_ACCENT = (231, 76, 60)        # Rouge pour montants
    COLOR_LIGHT_GRAY = (236, 240, 241)  # Gris clair pour arrière-plans
    COLOR_TEXT = (44, 62, 80)           # Texte principal
    
    def __init__(self):
        super().__init__()
        self.set_auto_page_break(auto=True, margin=15)
    
    def header_with_logo(self, logo_path=None, company_info=None):
        """En-tête professionnel avec logo et informations entreprise"""
        
        # Bande de couleur en haut
        self.set_fill_color(*self.COLOR_PRIMARY)
        self.rect(0, 0, 210, 5, 'F')
        
        # Logo à gauche
        if logo_path and os.path.exists(logo_path):
            try:
                optimized_logo = self.optimize_image(logo_path)
                if optimized_logo:
                    self.image(optimized_logo, x=10, y=10, w=50)
            except Exception as e:
                print(f"⚠️ Erreur logo : {e}")
        
        # Informations entreprise à droite
        if company_info:
            self.set_xy(120, 10)
            self.set_font("Helvetica", "B", 10)
            self.set_text_color(*self.COLOR_SECONDARY)
            
            # if 'company_name' in company_info:
            #     self.cell(0, 5, company_info['company_name'], new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
            if 'address' in company_info:
                self.set_font("Helvetica", "", 9)
                self.set_x(120)
                self.cell(0, 5, company_info['address'], new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
            if 'city' in company_info:
                self.set_x(120)
                self.cell(0, 5, company_info['city'], new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
            if 'phone' in company_info:
                self.set_x(120)
                self.cell(0, 5, f"Tel : {company_info['phone']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
            if 'email' in company_info:
                self.set_x(120)
                self.set_text_color(*self.COLOR_PRIMARY)
                self.cell(0, 5, company_info['email'], new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
        
        self.ln(10)
    
    def add_invoice_title(self, invoice_number, invoice_date):
        """Ajout des dates de facturation et de livraison l'une en dessous de l'autre"""
        self.ln(5)
        
        # Titre FACTURE
        self.set_fill_color(*self.COLOR_PRIMARY)
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 22)
        
        title = "FACTURE"
        title_width = 60
        self.set_x((210 - title_width) / 2)
        self.cell(title_width, 12, title, border=0, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C', fill=True)
        
        self.ln(5)
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(*self.COLOR_SECONDARY)
        
        # --- BLOC DATES (Aligné à droite) ---
        
        # 1. Ligne du Numéro et Date de Facturation
        self.set_x(15)
        self.cell(90, 6, f"N° Facture : {invoice_number}", align='L')
        # On place la date de facturation à droite
        self.cell(90, 6, f"Date de facturation : {invoice_date}", align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        # 2. Ligne Date de Livraison (Juste en dessous)
        # On ne change pas le X (car on a fait NEXT), on aligne juste à droite
        self.set_x(15) # On revient au bord gauche
        self.cell(180, 6, f"Date de livraison : {invoice_date}", align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        self.ln(5)
        
    def add_client_section(self, client_info):
        """Section client avec encadré"""
        self.set_fill_color(*self.COLOR_LIGHT_GRAY)
        self.rect(15, self.get_y(), 180, 35, 'F')
        
        # self.set_font("Helvetica", "B", 11)
        # self.set_text_color(*self.COLOR_PRIMARY)
        # self.set_x(20)
        # self.cell(0, 8, "FACTURE", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(*self.COLOR_TEXT)
        self.set_x(20)
        self.cell(0, 6, f"Acheteur : {client_info.get('company_name', '')}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        self.set_font("Helvetica", "", 9)
        self.set_x(20)
        self.cell(0, 5, client_info.get('postal_address', ''), new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        if 'siret' in client_info:
            self.set_x(20)
            self.set_font("Helvetica", "I", 9)
            self.cell(0, 5, f"SIREN : {client_info['siret']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        self.ln(10)
    
    def add_items_table(self, items):
        """Tableau avec colonne NATURE (Forfait)"""
        self.set_fill_color(*self.COLOR_PRIMARY)
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 9)
        
        # On ajuste les largeurs pour insérer "NATURE"
        # Total largeur = 180 (de 15 à 195)
        col_widths = [85, 25, 20, 25, 25] 
        headers = ["DESCRIPTION", "NATURE", "QTÉ", "P.U. HT", "TOTAL HT"]
        
        self.set_x(15)
        for i, header in enumerate(headers):
            self.cell(col_widths[i], 8, header, border=1, align='C', fill=True)
        self.ln()
        
        self.set_text_color(*self.COLOR_TEXT)
        self.set_font("Helvetica", "", 9)
        
        for item in items:
            self.set_x(15)
            # Description
            self.cell(col_widths[0], 8, str(item.get('description', '')[:45]), border=1)
            # Nature (Forfait en dur ou via data)
            self.cell(col_widths[1], 8, "Forfait", border=1, align='C')
            # Quantité
            self.cell(col_widths[2], 8, str(item.get('quantity', 1)), border=1, align='C')
            # Prix U
            self.cell(col_widths[3], 8, f"{item.get('unit_price', 0):.2f}", border=1, align='R')
            # Total
            self.cell(col_widths[4], 8, f"{item.get('total', 0):.2f}", border=1, align='R')
            self.ln()
        self.ln(5)
    
    def add_totals_section(self, subtotal, tax_rate, tax_amount, total):
        """Section totaux avec mise en valeur"""
        
        x_start = 130
        col_label = 40
        col_value = 30
        
        self.set_font("Helvetica", "", 10)
        self.set_text_color(*self.COLOR_TEXT)
        
        # Sous-total HT
        self.set_x(x_start)
        self.cell(col_label, 7, "Sous-total HT", align='R')
        self.cell(col_value, 7, f"{subtotal:.2f} EUR", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
        
        # TVA
        self.set_x(x_start)
        self.cell(col_label, 7, f"TVA ({tax_rate}%)", align='R')
        self.cell(col_value, 7, f"{tax_amount:.2f} EUR", new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R')
        
        # Ligne de séparation
        self.set_draw_color(*self.COLOR_PRIMARY)
        self.set_line_width(0.5)
        self.line(x_start, self.get_y() + 2, x_start + col_label + col_value, self.get_y() + 2)
        self.ln(5)
        
        # Total TTC
        self.set_fill_color(*self.COLOR_PRIMARY)
        self.set_text_color(255, 255, 255)
        self.set_font("Helvetica", "B", 12)
        
        self.set_x(x_start)
        self.cell(col_label, 10, "TOTAL TTC", border=1, new_x=XPos.RIGHT, new_y=YPos.TOP, align='R', fill=True)
        self.set_fill_color(*self.COLOR_ACCENT)
        self.cell(col_value, 10, f"{total:.2f} EUR", border=1, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='R', fill=True)
        
        self.ln(10)
    
    def add_payment_info(self, payment_info):
        """Informations de paiement"""
        self.set_font("Helvetica", "B", 10)
        self.set_text_color(*self.COLOR_PRIMARY)
        self.set_x(15)
        self.cell(0, 6, "INFORMATIONS DE REGLEMENT", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        self.set_font("Helvetica", "", 9)
        self.set_text_color(*self.COLOR_TEXT)
        
        if 'method' in payment_info:
            self.set_x(15)
            self.cell(0, 5, f"Mode de règlement : {payment_info['method']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        if 'due_date' in payment_info:
            self.set_x(15)
            self.cell(0, 5, f"Date d'écheance : {payment_info['due_date']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        if 'iban' in payment_info:
            self.set_x(15)
            self.cell(0, 5, f"IBAN : {payment_info['iban']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        if 'bic' in payment_info:
            self.set_x(15)
            self.cell(0, 5, f"BIC : {payment_info['bic']}", new_x=XPos.LMARGIN, new_y=YPos.NEXT)
        
        self.ln(5)
    
    def add_footer_legal(self, legal_texts):
        """Mentions légales en pied de page"""
        self.set_y(-40)
        
        self.set_draw_color(*self.COLOR_PRIMARY)
        self.set_line_width(0.3)
        self.line(15, self.get_y(), 195, self.get_y())
        self.ln(3)
        
        self.set_font("Helvetica", "I", 8)
        self.set_text_color(100, 100, 100)
        
        for text in legal_texts:
            self.cell(0, 4, text, new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
    
    @staticmethod
    def optimize_image(image_path, max_size=(800, 800)):
        """Optimise une image pour le PDF"""
        try:
            img = Image.open(image_path)
            
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if img.mode == 'RGBA' else None)
                img = background
            
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
            
            temp_path = "/tmp/optimized_logo.jpg"
            img.save(temp_path, "JPEG", quality=85, optimize=True)
            
            return temp_path
        except Exception as e:
            print(f"❌ Erreur optimisation image : {e}")
            return None


def generate_professional_invoice(data, output_file):
    """Génère une facture professionnelle (compatible fpdf2 moderne)"""
    
    try:
        if isinstance(data, str):
            data = json.loads(data)
        
        if VERBOSE == "yes":
            print("📄 Génération de la facture professionnelle...")
        
        pdf = ProfessionalInvoice()
        pdf.add_page()
        
        # 1. En-tête
        company_info = {
            'company_name': data.get('company_name', 'Votre Entreprise'),
            'address': data.get('billing_address', ''),
            'city': data.get('billing_city', ''),
            'email': data.get('issuer_email', '')
        }
        
        pdf.header_with_logo(
            logo_path=data.get('logo_path'),
            company_info=company_info
        )
        
        # 2. Titre
        pdf.add_invoice_title(
            invoice_number=data.get('invoice_number', 'XXXX'),
            invoice_date=data.get('Date', datetime.now().strftime('%d/%m/%Y'))
        )
        
        # 3. Client
        client_info = {
            'company_name': data.get('company_name', ''),
            'postal_address': data.get('postal_address', ''),
            'siret': data.get('Siret', '')
        }
        pdf.add_client_section(client_info)
        
        # 4. Prestations
        items = []
        
        if 'Prestation' in data:
            prix_ht_str = str(data.get('prix HT', '0'))
            prix_ht_clean = prix_ht_str.replace('€', '').replace('euros', '').replace('EUR', '').replace('eur', '').replace(',', '.').strip()
            try:
                prix_ht_value = float(prix_ht_clean)
            except ValueError:
                logger.exception(f"Erreur calcul prix_ht_value")
                prix_ht_value = 0.0
            
            items.append({
                'description': data['Prestation'],
                'quantity': 1,
                'unit_price': prix_ht_value,
                'total': prix_ht_value
            })
        
        if 'items' in data:
            items = data['items']
        
        pdf.add_items_table(items)
        
        # 5. Totaux
        def parse_price(price_str):
            if not price_str:
                return 0.0
            clean_str = str(price_str).replace('€', '').replace('euros', '').replace('EUR', '').replace('eur', '').replace(',', '.').strip()
            try:
                return float(clean_str)
            except ValueError:
                logger.exception(f"Calcul des prix")
                return 0.0
        
        prix_ht = parse_price(data.get('prix HT', '0'))
        tva_amount = parse_price(data.get('TVA 20%', '0'))
        prix_ttc = parse_price(data.get('Prix TTC', '0'))
        
        pdf.add_totals_section(
            subtotal=prix_ht,
            tax_rate=20,
            tax_amount=tva_amount,
            total=prix_ttc
        )
        
        # 6. Paiement
        payment_info = {
            'method': data.get('Réglement', ''),
            'due_date': data.get('Date du règlement', ''),
        }
        pdf.add_payment_info(payment_info)
        
        # 7. Mentions légales
        legal_texts = [
            data.get('legal_baseline 1', ''),
            data.get('legal_baseline 2', ''),
        ]
        legal_texts = [text for text in legal_texts if text]
        
        pdf.add_footer_legal(legal_texts)
        
        # Sauvegarder
        pdf.output(output_file)
        if VERBOSE == "yes":
            print(f"✅ Facture professionnelle créée : {output_file}")
        
        return True
        
    except Exception as e:
        
        if VERBOSE == "yes":
            print(f"❌ generate_professional_invoice : {e}")

        logger.exception(f"Generate_professional_invoice")

        traceback.print_exc()
        return False

def invoice_edition(stripe_id, conn):

    timestamp                       = datetime.now().strftime("%Y%m%d_%H%M%S")
    timestamp_letter                = datetime.now().strftime("%Y%m%d")

    try:
        cursor = conn.cursor()
        
        # Requête paramétrée pour la sécurité

        sql = "SELECT file_price, billing_post_code, billing_address, billing_full_name, company, siren, billing_city, customer_email  FROM billing_stripe where stripe_id = %s" 

        cursor.execute(sql, (str(stripe_id),))    

        result = cursor.fetchone()

        if result:
            # Les données sont dans un tuple, dans l'ordre du SELECT
            file_price_ht     = result[0]
            billing_post_code = result[1]
            billing_address   = result[2]
            billing_full_name = result[3]
            company_name      = result[4]
            siren             = result[5]
            billing_city      = result[6]
            customer_email    = result[7]

        else:
            logger.exception(f"Aucune donnée trouvée pour ce stripe_id")
            
            return False, None, None

        #print(f"✅ insert ID, criteria and file_link: {stripe_id}")
        # print(f"file_price_ht:{file_price_ht}")  
        # print(f"billing_post_code:{billing_post_code}")  
        # print(f"billing_address:{billing_address}")  
        # print(f"billing_full_name:{billing_full_name}")  
        # print(f"company_name:{company_name}")  
        # print(f"siren:{siren}")  
        # print(f"billing_city:{billing_city}")  
        # print(f"customer_email:{customer_email}")  


        invoice_file_name               = f"markethings_invoice_{timestamp}_{stripe_id}_{customer_email}"
        local_invoice_file_path         = f"{PATH_LOCAL_INVOICE_FILE}/{invoice_file_name}"
        invoice_remote_file_path        = f"{PATH_REMOTE_INVOICE_FILE}/{invoice_file_name}"


    except mysql.connector.Error as err:
        logger.exception(f"Erreur de connexion : invoice_edition")
    
    finally:
        cursor.close()

    invoice_number = get_invoice_number(stripe_id,conn)

    taux_tva = 0.20    # 20 %

    # Calculs
    tva_amount  = file_price_ht * taux_tva
    price_ttc   = file_price_ht + tva_amount
  
     # ✅ Créer directement le JSON avec f-string
    invoice_json = f"""
    {{
        "invoice_number": "STRIPE-{invoice_number}",
        "Date": "{timestamp_letter}",
        "Siret": "{siren or 'N/A'}",
        "Internal_stripe_id": "{stripe_id}",
        "company_name": "{company_name}",
        "postal_address": "{billing_address}, {billing_post_code} {billing_city}, France",
        "Prestation": "Fourniture d'un fichier de données sur mesure.",
        "Réglement": "Via carte bancaire en ligne sur Stripe via le site markethings.io",
        "Date du règlement": "{timestamp_letter}",
        "prix HT": "{file_price_ht:.2f} euros",
        "TVA 20%": "{tva_amount:.2f} euros",
        "Prix TTC": "{price_ttc:.2f} euros",
        "logo_path": "{PATH_LOGO_MARKETHINGS}",
        "legal_baseline 1": "Markethings SAS, au capital de 15 000 euros, immatriculée au RCS NANTERRE 839 513 827",
        "legal_baseline 2": "9 bis avenue JOFFRE 92250 La Garenne Colombes - markethings.io"
    }}
    """

    #status = generate_invoice(invoice_json, local_invoice_file_path)
    status = generate_professional_invoice(invoice_json, local_invoice_file_path)

    return status, invoice_remote_file_path, local_invoice_file_path

@app.route('/count_bot_v1', methods=['POST'])
@require_api_key_v1
@trafic_control
@query_control
@memory_guard

def get_companies_v1():
    if not request.is_json:
        return jsonify({'error': 'Content-Type must be application/json'}), 400

    if VERBOSE:
        print(request)
    criteria = request.get_json()
    criteria_dict = criteria.get("criteria", criteria)

    execution_block = criteria_dict.get("execution_mode", {})
    security_block = criteria_dict.get("security", {})
    if security_block:
        ip_address = security_block.get("ip_address")
    else:
        ip_address = "0.0.0.0"

    #print(f"execution_block:{execution_block}")

    # Step 2: Get the specific value (output_type)
    # We provide "count" as a second argument so it has a default value
    mode = execution_block.get("output_type", "count")

    #print(f"mode:{mode}")
    result_count_companies_logic = count_companies_logic(criteria)

    if mode == "count":
        try:

            return jsonify({
                "status": "success",
                **result_count_companies_logic
            }), 200

        except ValueError as e:
            return jsonify({"error get_companies_v1": str(e)}), 400

        except Exception as e:
            logger.exception("Internal server error get_companies_v1")
            return jsonify({
                'error': 'Internal server error',
                'message': str(e)
            }), 500

        return jsonify({
            "status": "success",
            **result_count_companies_logic
        }), 200

    elif mode == "display":

        #print(f"result:{result}")
            #print(f"result display:{result_count_companies_logic}")

            # On extrait la liste directement depuis le dictionnaire 'result'
        if isinstance(result_count_companies_logic, dict):
            # On récupère la liste d'entiers (on utilise results_legal ou results_semantic)
            siren_list_legal    = result_count_companies_logic.get('results_legal', [])
            siren_list_semantic = result_count_companies_logic.get('results_semantic', [])

        else:
            siren_list_legal    = []
            siren_list_semantic = []


        # Maintenant on appelle votre fonction SQL avec cette liste d'entiers
        result_legal    = get_company_info(siren_list_legal, "origine: codes NAF")
        result_semantic = get_company_info(siren_list_semantic, "origine: sémantique")

        # 2. On construit la réponse
        # Puisque result_legal est un dictionnaire, .get() fonctionne parfaitement
        return jsonify({
            "status": "success",
            "company_info_semantic": result_semantic,
            "company_info_legal": result_legal.get("data", []),
            "metadata": result_legal.get("metadata", {})
        }), 200

    elif mode == "big_file":

        #print(f"result:{result}")
        file_link = get_company_file(criteria, result_count_companies_logic,"")
        #print(f"company_info:{company_info}")

        return jsonify({
            "status": "success",
            "file_link": file_link
        }), 200
    else:
        logger.exception("Incorrect data return mode: get_companies_v1")
        return jsonify({
        'error': 'Incorrect data return mode',
        'message':  'Error server'
        }), 500
    


@app.route('/count_bot_v2', methods=['POST'])
@require_api_key_v2
@trafic_control
@query_control
@memory_guard


def get_companies_v2():

    if not request.is_json:
        return jsonify({'error': 'Content-Type must be application/json'}), 400

    criteria = request.get_json()

    try:
        result_count_companies_logic = count_companies_logic(criteria)
        criteria_dict = criteria.get("criteria", criteria)
        # Step 1: Get the inner dictionary (execution_mode)
        # doc try to be compatible with 
        
        execution_block = criteria_dict.get("execution_mode", {})
        #print(f"execution_block:{execution_block}")

        # Step 2: Get the specific value (output_type)
        # We provide "count" as a second argument so it has a default value
        mode = execution_block.get("output_type", "count")

        #print(f"mode:{mode}")
        if mode == "count":

            return jsonify({
                "status": "success",
                **result_count_companies_logic
            }), 200

        elif mode == "display":

            #print(f"result:{result}")
            #print(f"result display:{result_count_companies_logic}")

            # On extrait la liste directement depuis le dictionnaire 'result'
            if isinstance(result_count_companies_logic, dict):
                # On récupère la liste d'entiers (on utilise results_legal ou results_semantic)
                siren_list_legal    = result_count_companies_logic.get('results_legal', [])
                siren_list_semantic = result_count_companies_logic.get('results_semantic', [])

            else:
                siren_list_legal    = []
                siren_list_semantic = []

            # Maintenant on appelle votre fonction SQL avec cette liste d'entiers
            result_legal    = get_company_info(siren_list_legal, "origine codes NAF")
            result_semantic = get_company_info(siren_list_semantic, "origine: sémantique")


            # 2. On construit la réponse
            # Puisque result_legal est un dictionnaire, .get() fonctionne parfaitement
            return jsonify({
                "status": "success",
                "company_info_semantic": result_legal.get("data", []),
                "company_info_legal": result_legal.get("data", []),
                "metadata": result_legal.get("metadata", {})
            }), 200

        elif mode == "big_file":

            result = count_companies_logic(criteria)

            #print(f"result:{result}")
            file_link = get_company_file(criteria, result)

            #print(f"company_info:{company_info}")

            return jsonify({
                "status": "success",
                "file_link": file_link
            }), 200
        else:
            logger.exception("Incorrect data return mode")
            return jsonify({
            'error': 'Incorrect data return mode',
            'message': str(e)
        }), 500

       


    except ValueError as e:
        logger.exception("Internal server error: get_companies_v2")

        return jsonify({"error": str(e)}), 400

    except Exception as e:
        logger.exception("Internal server error: get_companies_v2")
        return jsonify({
            'error': 'Internal server error',
            'message': str(e)
        }), 500




    
@app.route('/check_siren_build_file_V1', methods=['POST'])
@require_api_key_v1

def check_company():
    # Récupération du JSON envoyé dans le corps de la requête
    data                = request.get_json()
    siren               = data['siren']
    criteria            = data['criteria']
    billing_post_code   = data['billing_post_code']
    billing_address     = data['billing_address']
    billing_full_name   = data['billing_full_name']
    file_price          = data['file_price']
    company             = data['company']
    billing_city        = data['billing_city']


    if not data:
        return jsonify({"error": "JSON body is missing"}), 400

    # Logique de vérification : On vérifie si le SIREN existe

    try:
        conn            = get_db_connection()
        cursor          = conn.cursor(dictionary=True)
        exists          = check_siren_in_db(siren, conn)
        if exists:
            new_stripe_id   = secrets.token_hex(10)
            code = 200

            # lancement du calcul du fichier
            result = count_companies_logic(criteria)
            # print("result")
            # print(result)
             # On nettoie l'email (remplacement des caractères spéciaux par des underscores)
            #clean_email = re.sub(r'[^a-zA-Z0-9]', '_', raw_email)
            
            # 2. Génération du Timestamp (format: 20240522_143005)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # 3. Construction du nom de fichier dynamique
            file_name           = f"export_{timestamp}_{new_stripe_id}.xlsx"
            #print(file_name)
            local_data_file     = f"./customer_files/data/{file_name}"
            data_remote_file_path    = f"{PATH_REMOTE_DATA_FILE}/{file_name}"

            print(f"*******************data_remote_file_path:{data_remote_file_path}")
            list_siren = result['siren_list']

            # 1. Préparer le thread
            thread_prod = threading.Thread(
                target=get_company_file, 
                args=(data_remote_file_path, local_data_file, criteria, list_siren, new_stripe_id)
            )

            # 2. Lancer le thread (ne bloque pas le script)
            thread_prod.start()

            # print("file_link")
            # print(file_link)
            # print(type(criteria))

            insert_stripe_id_file_link_criteria(new_stripe_id, file_price, billing_post_code, billing_address, billing_city, billing_full_name, local_data_file, data_remote_file_path, company, siren, criteria, conn)


        else:
            new_stripe_id = None
            logger.error(f"siren non existant: {siren}")
            code = 404

        # doc generation Id

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()

    # Vous pouvez ici ajouter une logique pour valider les 'criteria' 
    # envoyés dans le JSON si nécessaire.
    
    return jsonify({
        "siren": siren,
        "exists": exists,
        "stripe_id": new_stripe_id
    }), code


# doc webhook call when sucess payment
@app.route('/purchase_success_V1', methods=['POST'])
@require_api_key_v1

def purchase_success_V1():

     # Récupération du JSON envoyé dans le corps de la requête
    data            = request.get_json()
    stripe_id       = data['stripe_id']
    email_client    = data['email_client']
    card_owner      = data['card_owner']

    try:
        conn            = get_db_connection()
        cursor          = conn.cursor(dictionary=True)
        # doc update payment info and customer_email
        update_payment_info_email(stripe_id, email_client, card_owner, conn)

        local_data_file_path, data_remote_file_path = get_data_file_link(stripe_id, conn)
        data_remote_file_with_email_path       = data_remote_file_path.replace(".xlsx", f"_{email_client}.xlsx")

        # doc verification of the production od the data file
        status_file_ready  = wait_for_data_file(stripe_id, conn)

        if status_file_ready:
            # doc in voice production
            status, invoice_remote_file_path, invoice_local_file_path = invoice_edition(stripe_id, conn)
            update_payment_invoice(stripe_id, invoice_remote_file_path, conn)

        else:
             return jsonify({
                    "error":"data_file not build",
                    "success": False
                }), 404
    except Exception as e:
        logger.exception("Internal server error: purchase_success_V1")
        return jsonify({
            'error': 'Internal server error',
            'message': str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    if VERBOSE == "yes":
        print("**********************")
        print(f"local_data_file_path:{local_data_file_path}")
        print(f"data_remote_file_with_email_path:{data_remote_file_with_email_path}")
        print(f"invoice_local_file_path:{invoice_local_file_path}")
        print(f"invoice_remote_file_path:{invoice_remote_file_path}")


    status_push_delivery_files = push_delivery_files(local_data_file_path, data_remote_file_with_email_path, invoice_local_file_path, invoice_remote_file_path )

    data_remote_file_with_email_link = data_remote_file_with_email_path.replace(PATH_REMOTE_OVH, LINK_REMOTE_OVH)
    #print(f"data_remote_file_with_email_link:{data_remote_file_with_email_link}")
    invoice_remote_file_link = invoice_remote_file_path.replace(PATH_REMOTE_OVH, LINK_REMOTE_OVH)
    #print(f"data_remote_file_with_email_link:{data_remote_file_with_email_link}")
 
    return jsonify({
        "data_remote_file_path":data_remote_file_with_email_link,
        "invoice_remote_file_path":invoice_remote_file_link,
        "success": True
    }), 200


# doc déclencehemtnvia le webhook
@app.route('/get_info_success_page_V1', methods=['POST'])
@require_api_key_v1

def get_info_success_page():
     # Récupération du JSON envoyé dans le corps de la requête
    data            = request.get_json()
    stripe_id       = data['stripe_id']
  
    try:
        conn            = get_db_connection()
        cursor          = conn.cursor(dictionary=True)
        # doc update payment info and customer_email
        invoice_link_file, data_link_file = get_info_success_page(stripe_id, conn)
        # doc in voice production

    except Exception as e: 
        logger.exception("Internal server error get_info_success_page")
        return jsonify({
            'error': 'Internal server error',
            'message': str(e)
        }), 500

    finally:
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    
    return jsonify({
        "invoice_link_file":invoice_link_file,
        "data_link_file":data_link_file,
        "success": True
    }), 200



@app.route('/health', methods=['GET'])
def health_check():
    """Endpoint pour vérifier que l'API fonctionne"""
    return jsonify({
        'status': 'healthy',
        'timestamp': datetime.now().isoformat()
    }), 200


@app.errorhandler(404)
def not_found(error):
    return jsonify({
        'error': 'Endpoint not found'
    }), 404


@app.errorhandler(405)
def method_not_allowed(error):
    return jsonify({
        'error': 'Method not allowed'
    }), 405


if __name__ == '__main__':
    # Pour le développement
    app.run(host='0.0.0.0', port=5001, debug=True, threaded=True)

    # Pour la production, utilisez un serveur WSGI comme Gunicorn:
    # gunicorn -w 4 -b 0.0.0.0:5000 app:app